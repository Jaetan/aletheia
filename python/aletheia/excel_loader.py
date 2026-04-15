"""Excel loader for CAN signal checks and DBC definitions

Loads check definitions and DBC signal tables from Excel workbooks
(.xlsx) and compiles them through the Check API into LTL properties.

Designed for automotive technicians who define checks in spreadsheet
templates — no code, no YAML, just fill in cells.

Usage:
    from aletheia import load_checks_from_excel, load_dbc_from_excel

    # Load signal checks
    checks = load_checks_from_excel("my_checks.xlsx")
    client.set_properties([c.to_dict() for c in checks])

    # Load DBC definition
    dbc = load_dbc_from_excel("my_dbc.xlsx")
    client.parse_dbc(dbc)

    # Create a blank template for technicians
    create_template("template.xlsx")

Excel Template Layout
=====================

The workbook has three sheets:

**DBC** — one row per signal::

    Message ID | Message Name | DLC | Signal | Start Bit | Length |
    Byte Order | Signed | Factor | Offset | Min | Max | Unit |
    Multiplexor | Multiplex Value

Multiplexor and Multiplex Value are optional. If both are filled in, the
signal is multiplexed; if both are empty, the signal is always-present.

**Checks** — one row per simple check::

    Check Name | Signal | Condition | Value | Min | Max | Time (ms) | Severity

**When-Then** — one row per causal check::

    Check Name | When Signal | When Condition | When Value |
    Then Signal | Then Condition | Then Value | Then Min | Then Max |
    Within (ms) | Severity
"""

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .checks import Check, CheckResult
from ._check_conditions import (
    ALL_SIMPLE_CONDITIONS,
    SIMPLE_VALUE_CONDITIONS,
    SIMPLE_RANGE_CONDITIONS,
    SIMPLE_SETTLES_CONDITIONS,
    SIMPLE_EQUALS_CONDITIONS,
    WHEN_CONDITIONS,
    ALL_THEN_CONDITIONS,
    dispatch_simple,
    dispatch_when,
)
from .protocols import (
    DBCDefinition,
    DBCMessage,
    DBCSignal,
    DBCSignalAlways,
    DBCSignalMultiplexed,
)
from ._loader_utils import (
    is_str,
    get_str,
    get_number,
    get_int,
    get_bool,
)
from .client import to_signal_fraction

# Excel cell values: str, numbers, booleans, or None (empty)
# PEP 695 native ``type`` statement — lazy, no ``from __future__ import``
# needed, and basedpyright treats it as a proper type alias rather than a
# module-level variable.
type CellValue = str | int | float | bool | None

# A row of cell values as returned by openpyxl iter_rows(values_only=True)
type ExcelRow = tuple[CellValue, ...]


@dataclass(frozen=True)
class _MessageKey:
    """Grouping key for DBC message rows."""

    msg_id: int
    name: str
    extended: bool
    dlc: int


# ============================================================================
# Sheet headers
# ============================================================================

_DBC_HEADERS = [
    "Message ID", "Message Name", "Extended", "DLC", "Signal", "Start Bit",
    "Length", "Byte Order", "Signed", "Factor", "Offset", "Min", "Max",
    "Unit", "Multiplexor", "Multiplex Value",
]

_CHECKS_HEADERS = [
    "Check Name", "Signal", "Condition", "Value", "Min", "Max",
    "Time (ms)", "Severity",
]

_WHEN_THEN_HEADERS = [
    "Check Name", "When Signal", "When Condition", "When Value",
    "Then Signal", "Then Condition", "Then Value", "Then Min", "Then Max",
    "Within (ms)", "Severity",
]


# ============================================================================
# Row helpers
# ============================================================================

def _row_ctx(row_num: int) -> str:
    """Format a row number as an error context string."""
    return f"Row {row_num}"


def _row_to_dict(headers: list[str], row: ExcelRow) -> dict[str, object]:
    """Zip headers with cell values, skipping None values."""
    return {h: v for h, v in zip(headers, row, strict=False) if v is not None}


def _headers_from_row(row: ExcelRow) -> list[str]:
    """Extract header strings from the first row of a sheet."""
    return [str(c) if c is not None else "" for c in row]


def _parse_message_id(val: object, ctx: str) -> int:
    """Parse a message ID from an int or hex-string cell value."""
    if isinstance(val, int) and not isinstance(val, bool):
        return val
    if is_str(val):
        stripped = val.strip()
        try:
            if stripped.lower().startswith("0x"):
                return int(stripped, 16)
            return int(stripped)
        except ValueError:
            pass
    raise ValueError(
        f"{ctx}: invalid 'Message ID' — expected integer or hex string (e.g. 0x100)"
    )


# ============================================================================
# Public API
# ============================================================================

def load_checks_from_excel(
    path: str | Path,
    *,
    checks_sheet: str = "Checks",
    when_then_sheet: str = "When-Then",
) -> list[CheckResult]:
    """Load signal checks from an Excel workbook.

    Reads the Checks and When-Then sheets. Either or both may be present.

    Args:
        path: Path to a .xlsx workbook
        checks_sheet: Name of the simple-checks sheet
        when_then_sheet: Name of the when/then-checks sheet

    Returns:
        List of CheckResult objects ready for use with AletheiaClient

    Raises:
        FileNotFoundError: File doesn't exist
        ValueError: Invalid data in cells
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        results: list[CheckResult] = []
        sheet_names = wb.sheetnames

        if checks_sheet in sheet_names:
            results.extend(_load_simple_checks(wb[checks_sheet]))

        if when_then_sheet in sheet_names:
            results.extend(_load_when_then_checks(wb[when_then_sheet]))

        if checks_sheet not in sheet_names and when_then_sheet not in sheet_names:
            raise ValueError(
                f"Workbook has no '{checks_sheet}' or '{when_then_sheet}' sheet"
            )

        return results
    finally:
        wb.close()


def load_dbc_from_excel(
    path: str | Path,
    *,
    sheet: str = "DBC",
) -> DBCDefinition:
    """Load a DBC definition from the DBC sheet of an Excel workbook.

    Args:
        path: Path to a .xlsx workbook
        sheet: Name of the DBC sheet

    Returns:
        DBCDefinition dict ready for use with AletheiaClient.parse_dbc()

    Raises:
        FileNotFoundError: File doesn't exist
        ValueError: Invalid or missing data
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Workbook has no '{sheet}' sheet")

        ws = wb[sheet]
        rows: list[ExcelRow] = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise ValueError("DBC sheet must have a header row and at least one data row")

        headers = _headers_from_row(rows[0])
        data_rows = [_row_to_dict(headers, r) for r in rows[1:]]
        # Filter out completely empty rows
        data_rows = [r for r in data_rows if r]

        if not data_rows:
            raise ValueError("DBC sheet has no data rows")

        return _parse_dbc_rows(data_rows)
    finally:
        wb.close()


def create_template(path: str | Path) -> None:
    """Create a blank Excel template with headers and column formatting.

    Writes a .xlsx file with three sheets (DBC, Checks, When-Then),
    each with correct headers in bold. Does not overwrite existing files.

    Args:
        path: Output path for the .xlsx file

    Raises:
        FileExistsError: File already exists
    """
    p = Path(path)
    if p.exists():
        raise FileExistsError(f"File already exists: {path}")

    wb = Workbook()

    # DBC sheet (rename the default sheet)
    ws_dbc = wb.active
    if ws_dbc is None:
        raise RuntimeError("Workbook has no active sheet")
    ws_dbc.title = "DBC"
    _write_header_row(ws_dbc, _DBC_HEADERS)

    # Checks sheet
    ws_checks = wb.create_sheet("Checks")
    _write_header_row(ws_checks, _CHECKS_HEADERS)

    # When-Then sheet
    ws_wt = wb.create_sheet("When-Then")
    _write_header_row(ws_wt, _WHEN_THEN_HEADERS)

    wb.save(str(p))


# ============================================================================
# Internal: sheet loaders
# ============================================================================

def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write bold header row to a worksheet."""
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold


def _load_simple_checks(ws: Worksheet) -> list[CheckResult]:
    """Load CheckResults from a Checks sheet."""
    rows: list[ExcelRow] = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = _headers_from_row(rows[0])
    results: list[CheckResult] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        d = _row_to_dict(headers, row)
        if not d:
            continue  # skip empty rows
        results.append(_parse_simple_row(d, row_idx))
    return results


def _load_when_then_checks(ws: Worksheet) -> list[CheckResult]:
    """Load CheckResults from a When-Then sheet."""
    rows: list[ExcelRow] = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = _headers_from_row(rows[0])
    results: list[CheckResult] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        d = _row_to_dict(headers, row)
        if not d:
            continue  # skip empty rows
        results.append(_parse_when_then_row(d, row_idx))
    return results


# ============================================================================
# Internal: row parsers
# ============================================================================

def _apply_metadata(result: CheckResult, d: dict[str, object]) -> CheckResult:
    """Apply optional name and severity from row data to a CheckResult."""
    name = d.get("Check Name")
    if is_str(name):
        result.named(name)
    sev = d.get("Severity")
    if is_str(sev):
        result.severity(sev)
    return result


def _parse_simple_row(d: dict[str, object], row_num: int) -> CheckResult:
    """Parse a Checks-sheet row into a CheckResult."""
    signal = get_str(d, "Signal", _row_ctx(row_num))
    condition = get_str(d, "Condition", _row_ctx(row_num))

    if condition not in ALL_SIMPLE_CONDITIONS:
        raise ValueError(f"Row {row_num}: unknown condition '{condition}'")

    if condition in SIMPLE_VALUE_CONDITIONS:
        result = dispatch_simple(signal, condition, get_number(d, "Value", _row_ctx(row_num)))
    elif condition in SIMPLE_RANGE_CONDITIONS:
        if "Min" not in d or "Max" not in d:
            raise ValueError(
                f"Row {row_num}: condition '{condition}' requires 'Min' and 'Max'"
            )
        result = Check.signal(signal).stays_between(
            get_number(d, "Min", _row_ctx(row_num)),
            get_number(d, "Max", _row_ctx(row_num)),
        )
    elif condition in SIMPLE_SETTLES_CONDITIONS:
        if "Min" not in d or "Max" not in d:
            raise ValueError(
                f"Row {row_num}: condition 'settles_between' requires 'Min' and 'Max'"
            )
        if "Time (ms)" not in d:
            raise ValueError(
                f"Row {row_num}: condition 'settles_between' requires 'Time (ms)'"
            )
        result = Check.signal(signal).settles_between(
            get_number(d, "Min", _row_ctx(row_num)),
            get_number(d, "Max", _row_ctx(row_num)),
        ).within(get_int(d, "Time (ms)", _row_ctx(row_num)))
    elif condition in SIMPLE_EQUALS_CONDITIONS:
        value = get_number(d, "Value", _row_ctx(row_num))
        result = Check.signal(signal).equals(value).always()
    else:
        raise ValueError(f"Row {row_num}: unknown condition '{condition}'")

    return _apply_metadata(result, d)


def _parse_when_then_row(d: dict[str, object], row_num: int) -> CheckResult:
    """Parse a When-Then-sheet row into a CheckResult."""
    # When clause
    when_signal = get_str(d, "When Signal", _row_ctx(row_num))
    when_cond = get_str(d, "When Condition", _row_ctx(row_num))
    when_value = get_number(d, "When Value", _row_ctx(row_num))

    if when_cond not in WHEN_CONDITIONS:
        raise ValueError(f"Row {row_num}: unknown when condition '{when_cond}'")

    when_result = dispatch_when(Check.when(when_signal), when_cond, when_value)

    # Then clause
    then_signal = get_str(d, "Then Signal", _row_ctx(row_num))
    then_cond = get_str(d, "Then Condition", _row_ctx(row_num))

    if then_cond not in ALL_THEN_CONDITIONS:
        raise ValueError(f"Row {row_num}: unknown then condition '{then_cond}'")

    then_builder = when_result.then(then_signal)

    if then_cond == "equals":
        then_result = then_builder.equals(get_number(d, "Then Value", _row_ctx(row_num)))
    elif then_cond == "exceeds":
        then_result = then_builder.exceeds(get_number(d, "Then Value", _row_ctx(row_num)))
    else:  # stays_between
        if "Then Min" not in d or "Then Max" not in d:
            raise ValueError(
                f"Row {row_num}: then condition 'stays_between' requires 'Then Min' and 'Then Max'"
            )
        then_result = then_builder.stays_between(
            get_number(d, "Then Min", _row_ctx(row_num)),
            get_number(d, "Then Max", _row_ctx(row_num)),
        )

    result = then_result.within(get_int(d, "Within (ms)", _row_ctx(row_num)))
    return _apply_metadata(result, d)


# ============================================================================
# Internal: DBC parser
# ============================================================================

def _parse_dbc_signal(row: dict[str, object], row_num: int) -> DBCSignal:
    """Parse a single DBC signal row into a DBCSignal dict."""
    byte_order = get_str(row, "Byte Order", _row_ctx(row_num))
    if byte_order not in ("little_endian", "big_endian"):
        raise ValueError(
            f"Row {row_num}: 'Byte Order' must be 'little_endian' or 'big_endian'"
        )

    unit = row.get("Unit")
    unit_str = unit if is_str(unit) else ""
    ctx = _row_ctx(row_num)

    has_muxor = "Multiplexor" in row
    has_mux_val = "Multiplex Value" in row

    if has_muxor != has_mux_val:
        raise ValueError(
            f"Row {row_num}: 'Multiplexor' and 'Multiplex Value' "
            + "must both be provided or both be empty"
        )

    if has_muxor:
        mux_signal: DBCSignalMultiplexed = {
            "name": get_str(row, "Signal", ctx),
            "startBit": get_int(row, "Start Bit", ctx),
            "length": get_int(row, "Length", ctx),
            "byteOrder": byte_order,
            "signed": get_bool(row, "Signed", ctx),
            "factor": to_signal_fraction(get_number(row, "Factor", ctx)),
            "offset": to_signal_fraction(get_number(row, "Offset", ctx)),
            "minimum": to_signal_fraction(get_number(row, "Min", ctx)),
            "maximum": to_signal_fraction(get_number(row, "Max", ctx)),
            "unit": unit_str,
            "multiplexor": get_str(row, "Multiplexor", ctx),
            "multiplex_values": [get_int(row, "Multiplex Value", ctx)],
        }
        return mux_signal

    always_signal: DBCSignalAlways = {
        "name": get_str(row, "Signal", ctx),
        "startBit": get_int(row, "Start Bit", ctx),
        "length": get_int(row, "Length", ctx),
        "byteOrder": byte_order,
        "signed": get_bool(row, "Signed", ctx),
        "factor": to_signal_fraction(get_number(row, "Factor", ctx)),
        "offset": to_signal_fraction(get_number(row, "Offset", ctx)),
        "minimum": to_signal_fraction(get_number(row, "Min", ctx)),
        "maximum": to_signal_fraction(get_number(row, "Max", ctx)),
        "unit": unit_str,
        "presence": "always",
    }
    return always_signal


def _parse_dbc_rows(rows: list[dict[str, object]]) -> DBCDefinition:
    """Group DBC rows by message and build a DBCDefinition."""
    groups: dict[_MessageKey, list[int]] = defaultdict(list)
    insertion_order: list[_MessageKey] = []

    for idx, row in enumerate(rows):
        row_num = idx + 2  # 1-indexed, skip header
        ext_val = row.get("Extended")
        is_extended = get_bool(row, "Extended", _row_ctx(row_num)) if ext_val is not None else False
        key = _MessageKey(
            msg_id=_parse_message_id(row.get("Message ID"), _row_ctx(row_num)),
            name=get_str(row, "Message Name", _row_ctx(row_num)),
            extended=is_extended,
            dlc=get_int(row, "DLC", _row_ctx(row_num)),
        )
        if key not in groups:
            insertion_order.append(key)
        groups[key].append(idx)

    messages: list[DBCMessage] = []
    for key in insertion_order:
        signals: list[DBCSignal] = [
            _parse_dbc_signal(rows[i], i + 2) for i in groups[key]
        ]
        msg = DBCMessage(
            id=key.msg_id,
            name=key.name,
            dlc=key.dlc,
            sender="",
            signals=signals,
        )
        if key.extended:
            msg["extended"] = True
        messages.append(msg)

    return {"version": "", "messages": messages}
