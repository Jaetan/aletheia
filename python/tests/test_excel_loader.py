"""Unit tests for the Excel loader.

Tests cover:
- Simple checks: each condition type loaded from Excel rows
- When/Then checks: all trigger/response combinations
- Metadata: name and severity from cells
- DBC loading: parse DBC sheet, message grouping, hex IDs
- Template creation: sheet names, headers, no-overwrite
- Error handling: invalid conditions, missing cells, bad message IDs
- File round-trip: write temp .xlsx, load it, verify
- Equivalence: Excel-loaded checks produce same formula as Check API
"""

from __future__ import annotations

from typing import TYPE_CHECKING, cast

import openpyxl  # type: ignore[import-untyped]
import pytest
from _excel_helpers import (
    ALWAYS_SIGNAL_MUX_COLS_ROW,
    BRAKE_PRESSURE_ROW,
    BRAKE_RESPONSE_SAFETY_WT_ROW,
    BRAKE_RESPONSE_WT_ROW,
    ENGINE_RPM_HEX_ID_ROW,
    ENGINE_RPM_NO_UNIT_ROW,
    ENGINE_RPM_ROW,
    ENGINE_TEMP_ROW,
    ENGINE_TEMP_SIGNED_ROW,
    FUEL_WT_ROW,
    IGNITION_RPM_WT_ROW,
    INVALID_BYTE_ORDER_ROW,
    INVALID_MESSAGE_ID_ROW,
    MIXED_SELECTOR_ROW,
    MIXED_TEMPA_ROW,
    MIXED_TEMPB_ROW,
    MUX_SIGNAL_ROW,
    PARTIAL_MUX_ROW,
    PARTIAL_MUX_VALUE_ROW,
    SIGNED_INT_ONE_ROW,
    SIGNED_INT_ZERO_ROW,
    SYMLINK_DBC_ROW,
    UNKNOWN_THEN_WT_ROW,
    UNKNOWN_WHEN_WT_ROW,
    active_sheet,
    make_checks_workbook,
    make_dbc_workbook,
    make_when_then_workbook,
)

from aletheia import ValidationError
from aletheia.checks import signal, when
from aletheia.excel_loader import (
    CHECKS_HEADERS,
    DBC_HEADERS,
    WHEN_THEN_HEADERS,
    create_template,
    load_checks_from_excel,
    load_dbc_from_excel,
)

if TYPE_CHECKING:
    from pathlib import Path

    from aletheia.protocols import DBCSignalAlways, DBCSignalMultiplexed


# ============================================================================
# Simple checks — each condition type
# ============================================================================


class TestLoadSimpleChecks:
    """Load each simple condition from Excel rows and verify formula."""

    def test_never_exceeds(self, tmp_path: Path) -> None:
        """Verify never exceeds."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Speed", "never_exceeds", 220, None, None, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == signal("Speed").never_exceeds(220).to_dict()

    def test_never_below(self, tmp_path: Path) -> None:
        """Verify never below."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Voltage", "never_below", 11.5, None, None, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == signal("Voltage").never_below(11.5).to_dict()

    def test_stays_between(self, tmp_path: Path) -> None:
        """Verify stays between."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Voltage", "stays_between", None, 11.5, 14.5, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == (signal("Voltage").stays_between(11.5, 14.5).to_dict())

    def test_never_equals(self, tmp_path: Path) -> None:
        """Verify never equals."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "ErrorCode", "never_equals", 255, None, None, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == signal("ErrorCode").never_equals(255).to_dict()

    def test_equals_always(self, tmp_path: Path) -> None:
        """Verify equals always."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Gear", "equals", 0, None, None, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == signal("Gear").equals(0).always().to_dict()

    def test_settles_between(self, tmp_path: Path) -> None:
        """Verify settles between."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "CoolantTemp", "settles_between", None, 80, 100, 5000, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == (
            signal("CoolantTemp").settles_between(80, 100).within(5000).to_dict()
        )

    def test_multiple_checks(self, tmp_path: Path) -> None:
        """Verify multiple checks."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Speed", "never_exceeds", 220, None, None, None, None],
                [None, "Voltage", "stays_between", None, 11.5, 14.5, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 2
        assert checks[0].to_dict() == signal("Speed").never_exceeds(220).to_dict()
        assert checks[1].to_dict() == (signal("Voltage").stays_between(11.5, 14.5).to_dict())


# ============================================================================
# When/Then checks
# ============================================================================


class TestLoadWhenThenChecks:
    """Load when/then causal checks from Excel rows."""

    def test_when_exceeds_then_equals(self, tmp_path: Path) -> None:
        """Verify when exceeds then equals."""
        p = make_when_then_workbook(tmp_path, [BRAKE_RESPONSE_WT_ROW])
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        expected = when("BrakePedal").exceeds(50).then("BrakeLight").equals(1).within(100).to_dict()
        assert checks[0].to_dict() == expected

    def test_when_equals_then_exceeds(self, tmp_path: Path) -> None:
        """Verify when equals then exceeds."""
        p = make_when_then_workbook(tmp_path, [IGNITION_RPM_WT_ROW])
        checks = load_checks_from_excel(p)
        expected = when("Ignition").equals(1).then("RPM").exceeds(500).within(2000).to_dict()
        assert checks[0].to_dict() == expected

    def test_when_drops_below_then_stays_between(self, tmp_path: Path) -> None:
        """Verify when drops below then stays between."""
        p = make_when_then_workbook(tmp_path, [FUEL_WT_ROW])
        checks = load_checks_from_excel(p)
        expected = (
            when("FuelLevel")
            .drops_below(10)
            .then("FuelWarning")
            .stays_between(1, 1)
            .within(50)
            .to_dict()
        )
        assert checks[0].to_dict() == expected


# ============================================================================
# Metadata
# ============================================================================


class TestLoadMetadata:
    """Verify name and severity are set on CheckResult from Excel cells."""

    def test_name_set(self, tmp_path: Path) -> None:
        """Verify name set."""
        p = make_checks_workbook(
            tmp_path,
            [
                ["Speed limit", "Speed", "never_exceeds", 220, None, None, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert checks[0].name == "Speed limit"

    def test_severity_set(self, tmp_path: Path) -> None:
        """Verify severity set."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Speed", "never_exceeds", 220, None, None, None, "critical"],
            ],
        )
        checks = load_checks_from_excel(p)
        assert checks[0].check_severity == "critical"

    def test_name_and_severity(self, tmp_path: Path) -> None:
        """Verify name and severity."""
        p = make_checks_workbook(
            tmp_path,
            [
                ["Speed limit", "Speed", "never_exceeds", 220, None, None, None, "warning"],
            ],
        )
        checks = load_checks_from_excel(p)
        assert checks[0].name == "Speed limit"
        assert checks[0].check_severity == "warning"

    def test_defaults_none(self, tmp_path: Path) -> None:
        """Verify defaults none."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Speed", "never_exceeds", 220, None, None, None, None],
            ],
        )
        checks = load_checks_from_excel(p)
        assert checks[0].name == ""
        assert checks[0].check_severity == ""

    def test_when_then_metadata(self, tmp_path: Path) -> None:
        """Verify when then metadata."""
        p = make_when_then_workbook(tmp_path, [BRAKE_RESPONSE_SAFETY_WT_ROW])
        checks = load_checks_from_excel(p)
        assert checks[0].name == "Brake response"
        assert checks[0].check_severity == "safety"


# ============================================================================
# DBC loading
# ============================================================================


class TestLoadDBCFromExcel:
    """Parse DBC sheet, verify DBCDefinition structure."""

    def test_single_signal(self, tmp_path: Path) -> None:
        """Verify single signal."""
        p = make_dbc_workbook(tmp_path, [ENGINE_RPM_ROW])
        dbc = load_dbc_from_excel(p)
        assert dbc["version"] == ""
        assert len(dbc["messages"]) == 1
        msg = dbc["messages"][0]
        assert msg["id"] == 256
        assert msg["name"] == "EngineData"
        assert msg["dlc"] == 8
        assert len(msg["signals"]) == 1
        # Fixture omits Multiplexor / Multiplex Value columns → always-present.
        # cast narrows the DBCSignalAlways | DBCSignalMultiplexed union so
        # presence-key access type-checks without a `# type: ignore`.
        sig = cast("DBCSignalAlways", msg["signals"][0])
        assert sig["name"] == "RPM"
        assert sig["startBit"] == 0
        assert sig["length"] == 16
        assert sig["byteOrder"] == "little_endian"
        assert sig["signed"] is False
        assert sig["factor"] == 0.25
        assert sig["offset"] == 0
        assert sig["minimum"] == 0
        assert sig["maximum"] == 16383.75
        assert sig["unit"] == "rpm"
        assert sig["presence"] == "always"

    def test_message_grouping(self, tmp_path: Path) -> None:
        """Multiple rows with same message ID are grouped."""
        p = make_dbc_workbook(
            tmp_path,
            [ENGINE_RPM_ROW, ENGINE_TEMP_ROW, BRAKE_PRESSURE_ROW],
        )
        dbc = load_dbc_from_excel(p)
        assert len(dbc["messages"]) == 2
        assert dbc["messages"][0]["name"] == "EngineData"
        assert len(dbc["messages"][0]["signals"]) == 2
        assert dbc["messages"][0]["signals"][0]["name"] == "RPM"
        assert dbc["messages"][0]["signals"][1]["name"] == "Temp"
        assert dbc["messages"][1]["name"] == "BrakeData"
        assert len(dbc["messages"][1]["signals"]) == 1

    def test_hex_message_id(self, tmp_path: Path) -> None:
        """Message IDs can be hex strings."""
        p = make_dbc_workbook(tmp_path, [ENGINE_RPM_HEX_ID_ROW])
        dbc = load_dbc_from_excel(p)
        assert dbc["messages"][0]["id"] == 0x100

    def test_signed_true(self, tmp_path: Path) -> None:
        """Verify signed true."""
        p = make_dbc_workbook(tmp_path, [ENGINE_TEMP_SIGNED_ROW])
        dbc = load_dbc_from_excel(p)
        assert dbc["messages"][0]["signals"][0]["signed"] is True

    def test_signed_integer_one(self, tmp_path: Path) -> None:
        """Signed column as integer 1 (Excel data_only mode) accepted."""
        p = make_dbc_workbook(tmp_path, [SIGNED_INT_ONE_ROW])
        dbc = load_dbc_from_excel(p)
        assert dbc["messages"][0]["signals"][0]["signed"] is True

    def test_signed_integer_zero(self, tmp_path: Path) -> None:
        """Signed column as integer 0 (Excel data_only mode) accepted."""
        p = make_dbc_workbook(tmp_path, [SIGNED_INT_ZERO_ROW])
        dbc = load_dbc_from_excel(p)
        assert dbc["messages"][0]["signals"][0]["signed"] is False

    def test_missing_unit_defaults_empty(self, tmp_path: Path) -> None:
        """Verify missing unit defaults empty."""
        p = make_dbc_workbook(tmp_path, [ENGINE_RPM_NO_UNIT_ROW])
        dbc = load_dbc_from_excel(p)
        assert dbc["messages"][0]["signals"][0]["unit"] == ""


# ============================================================================
# Multiplexed signals in DBC
# ============================================================================


class TestLoadDBCMultiplexed:
    """Multiplexed signal support via optional Multiplexor/Multiplex Value columns."""

    def test_multiplexed_signal(self, tmp_path: Path) -> None:
        """Signal with both Multiplexor and Multiplex Value produces DBCSignalMultiplexed."""
        p = make_dbc_workbook(tmp_path, [MUX_SIGNAL_ROW])
        dbc = load_dbc_from_excel(p)
        sig = dbc["messages"][0]["signals"][0]
        assert "multiplexor" in sig
        assert sig["multiplexor"] == "Selector"
        assert sig["multiplex_values"] == [3]
        # Multiplexed signals carry an explicit
        # ``"presence": "multiplexed"`` discriminator.
        assert sig["presence"] == "multiplexed"

    def test_always_signal_no_mux_columns(self, tmp_path: Path) -> None:
        """Signal without Multiplexor/Multiplex Value columns is always-present."""
        p = make_dbc_workbook(tmp_path, [ALWAYS_SIGNAL_MUX_COLS_ROW])
        dbc = load_dbc_from_excel(p)
        # cast narrows the DBCSignal union; the `multiplexor` key absence is
        # the structural invariant the test asserts.
        sig = cast("DBCSignalAlways", dbc["messages"][0]["signals"][0])
        assert sig["presence"] == "always"
        assert sig.get("multiplexor") is None

    def test_mixed_always_and_multiplexed(self, tmp_path: Path) -> None:
        """Same message can have both always-present and multiplexed signals."""
        p = make_dbc_workbook(
            tmp_path,
            [MIXED_SELECTOR_ROW, MIXED_TEMPA_ROW, MIXED_TEMPB_ROW],
        )
        dbc = load_dbc_from_excel(p)
        msg = dbc["messages"][0]
        assert len(msg["signals"]) == 3
        # First fixture row has no Multiplexor/Value columns → always-present.
        # Other two carry "Selector" / value 0|1 → multiplexed variant.
        sig0 = cast("DBCSignalAlways", msg["signals"][0])
        sig1 = cast("DBCSignalMultiplexed", msg["signals"][1])
        sig2 = cast("DBCSignalMultiplexed", msg["signals"][2])
        assert sig0["presence"] == "always"
        assert sig1["multiplexor"] == "Selector"
        assert sig1["multiplex_values"] == [0]
        assert sig2["multiplexor"] == "Selector"
        assert sig2["multiplex_values"] == [1]

    def test_partial_mux_raises(self, tmp_path: Path) -> None:
        """Only Multiplexor without Multiplex Value raises ValidationError."""
        p = make_dbc_workbook(tmp_path, [PARTIAL_MUX_ROW])
        with pytest.raises(ValidationError, match="must both be provided or both be empty"):
            load_dbc_from_excel(p)

    def test_partial_mux_value_only_raises(self, tmp_path: Path) -> None:
        """Only Multiplex Value without Multiplexor raises ValidationError."""
        p = make_dbc_workbook(tmp_path, [PARTIAL_MUX_VALUE_ROW])
        with pytest.raises(ValidationError, match="must both be provided or both be empty"):
            load_dbc_from_excel(p)


# ============================================================================
# Template creation
# ============================================================================


class TestCreateTemplate:
    """Verify create_template() creates a correct workbook."""

    def test_creates_file(self, tmp_path: Path) -> None:
        """Verify creates file."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        assert p.exists()

    def test_sheet_names(self, tmp_path: Path) -> None:
        """Verify sheet names."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        wb = openpyxl.load_workbook(p)
        assert wb.sheetnames == ["DBC", "Checks", "When-Then"]
        wb.close()

    def test_dbc_headers(self, tmp_path: Path) -> None:
        """Verify dbc headers."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        wb = openpyxl.load_workbook(p)
        ws = wb["DBC"]
        headers = [cell.value for cell in ws[1]]
        assert headers == DBC_HEADERS
        wb.close()

    def test_checks_headers(self, tmp_path: Path) -> None:
        """Verify checks headers."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        wb = openpyxl.load_workbook(p)
        ws = wb["Checks"]
        headers = [cell.value for cell in ws[1]]
        assert headers == CHECKS_HEADERS
        wb.close()

    def test_when_then_headers(self, tmp_path: Path) -> None:
        """Verify when then headers."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        wb = openpyxl.load_workbook(p)
        ws = wb["When-Then"]
        headers = [cell.value for cell in ws[1]]
        assert headers == WHEN_THEN_HEADERS
        wb.close()

    def test_headers_bold(self, tmp_path: Path) -> None:
        """Verify headers bold."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        wb = openpyxl.load_workbook(p)
        ws = wb["DBC"]
        for cell in ws[1]:
            assert cell.font.bold is True
        wb.close()

    def test_no_overwrite(self, tmp_path: Path) -> None:
        """Verify no overwrite."""
        p = tmp_path / "template.xlsx"
        create_template(p)
        with pytest.raises(FileExistsError, match="File already exists"):
            create_template(p)


# ============================================================================
# Error handling
# ============================================================================


class TestLoadErrors:
    """All validation error cases raise ValidationError or FileNotFoundError."""

    def test_file_not_found(self) -> None:
        """Verify file not found."""
        with pytest.raises(FileNotFoundError, match="Excel file not found"):
            load_checks_from_excel("/nonexistent/path/checks.xlsx")

    def test_dbc_file_not_found(self) -> None:
        """Verify dbc file not found."""
        with pytest.raises(FileNotFoundError, match="Excel file not found"):
            load_dbc_from_excel("/nonexistent/path/dbc.xlsx")

    def test_no_checks_or_when_then_sheet(self, tmp_path: Path) -> None:
        """Workbook with neither Checks nor When-Then raises ValidationError."""
        wb = openpyxl.Workbook()
        ws = active_sheet(wb)
        ws.title = "Other"
        p = tmp_path / "bad.xlsx"
        wb.save(str(p))
        with pytest.raises(ValidationError, match="no 'Checks' or 'When-Then' sheet"):
            load_checks_from_excel(p)

    def test_no_dbc_sheet(self, tmp_path: Path) -> None:
        """Verify no dbc sheet."""
        wb = openpyxl.Workbook()
        ws = active_sheet(wb)
        ws.title = "Other"
        p = tmp_path / "bad.xlsx"
        wb.save(str(p))
        with pytest.raises(ValidationError, match="no 'DBC' sheet"):
            load_dbc_from_excel(p)

    def test_unknown_simple_condition(self, tmp_path: Path) -> None:
        """Verify unknown simple condition."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Speed", "bogus", 100, None, None, None, None],
            ],
        )
        with pytest.raises(ValidationError, match="unknown condition 'bogus'"):
            load_checks_from_excel(p)

    def test_missing_value_for_never_exceeds(self, tmp_path: Path) -> None:
        """Verify missing value for never exceeds."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Speed", "never_exceeds", None, None, None, None, None],
            ],
        )
        with pytest.raises(ValidationError, match="missing or invalid 'Value'"):
            load_checks_from_excel(p)

    def test_stays_between_missing_min(self, tmp_path: Path) -> None:
        """Verify stays between missing min."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Voltage", "stays_between", None, None, 14.5, None, None],
            ],
        )
        with pytest.raises(ValidationError, match="requires 'Min' and 'Max'"):
            load_checks_from_excel(p)

    def test_settles_between_missing_time(self, tmp_path: Path) -> None:
        """Verify settles between missing time."""
        p = make_checks_workbook(
            tmp_path,
            [
                [None, "Temp", "settles_between", None, 80, 100, None, None],
            ],
        )
        with pytest.raises(ValidationError, match="requires 'Time \\(ms\\)'"):
            load_checks_from_excel(p)

    def test_unknown_when_condition(self, tmp_path: Path) -> None:
        """Verify unknown when condition."""
        p = make_when_then_workbook(tmp_path, [UNKNOWN_WHEN_WT_ROW])
        with pytest.raises(ValidationError, match="unknown when condition 'bogus'"):
            load_checks_from_excel(p)

    def test_unknown_then_condition(self, tmp_path: Path) -> None:
        """Verify unknown then condition."""
        p = make_when_then_workbook(tmp_path, [UNKNOWN_THEN_WT_ROW])
        with pytest.raises(ValidationError, match="unknown then condition 'bogus'"):
            load_checks_from_excel(p)

    def test_invalid_byte_order(self, tmp_path: Path) -> None:
        """Verify invalid byte order."""
        p = make_dbc_workbook(tmp_path, [INVALID_BYTE_ORDER_ROW])
        with pytest.raises(ValidationError, match="Byte Order"):
            load_dbc_from_excel(p)

    def test_invalid_message_id(self, tmp_path: Path) -> None:
        """Verify invalid message id."""
        p = make_dbc_workbook(tmp_path, [INVALID_MESSAGE_ID_ROW])
        with pytest.raises(ValidationError, match="invalid 'Message ID'"):
            load_dbc_from_excel(p)

    def test_dbc_empty_data(self, tmp_path: Path) -> None:
        """DBC sheet with only header row raises ValidationError."""
        wb = openpyxl.Workbook()
        ws = active_sheet(wb)
        ws.title = "DBC"
        ws.append(DBC_HEADERS)
        p = tmp_path / "empty.xlsx"
        wb.save(str(p))
        with pytest.raises(ValidationError, match="at least one data row"):
            load_dbc_from_excel(p)


# ============================================================================
# File round-trip
# ============================================================================


class TestLoadFromFile:
    """Write temp .xlsx, load it, verify round-trip."""

    def test_checks_round_trip(self, tmp_path: Path) -> None:
        """Verify checks round trip."""
        p = make_checks_workbook(
            tmp_path,
            [
                ["Speed limit", "Speed", "never_exceeds", 220, None, None, None, "critical"],
            ],
        )
        checks = load_checks_from_excel(p)
        assert len(checks) == 1
        assert checks[0].to_dict() == signal("Speed").never_exceeds(220).to_dict()
        assert checks[0].name == "Speed limit"
        assert checks[0].check_severity == "critical"

    def test_combined_checks_and_when_then(self, tmp_path: Path) -> None:
        """Workbook with both Checks and When-Then sheets."""
        wb = openpyxl.Workbook()
        ws_checks = active_sheet(wb)
        ws_checks.title = "Checks"
        ws_checks.append(CHECKS_HEADERS)
        ws_checks.append(
            [None, "Speed", "never_exceeds", 220, None, None, None, None],
        )

        ws_wt = wb.create_sheet("When-Then")
        ws_wt.append(WHEN_THEN_HEADERS)
        ws_wt.append(
            [None, "Brake", "exceeds", 50, "BrakeLight", "equals", 1, None, None, 100, None],
        )

        p = tmp_path / "combined.xlsx"
        wb.save(str(p))

        checks = load_checks_from_excel(p)
        assert len(checks) == 2
        # First is the simple check, second is the when/then
        assert checks[0].to_dict() == signal("Speed").never_exceeds(220).to_dict()
        expected_wt = when("Brake").exceeds(50).then("BrakeLight").equals(1).within(100).to_dict()
        assert checks[1].to_dict() == expected_wt

    def test_dbc_round_trip(self, tmp_path: Path) -> None:
        """Verify dbc round trip."""
        p = make_dbc_workbook(tmp_path, [ENGINE_RPM_ROW, ENGINE_TEMP_ROW])
        dbc = load_dbc_from_excel(p)
        assert len(dbc["messages"]) == 1
        msg = dbc["messages"][0]
        assert msg["id"] == 256
        assert len(msg["signals"]) == 2
        assert msg["signals"][0]["name"] == "RPM"
        assert msg["signals"][1]["name"] == "Temp"


# ============================================================================
# Skips empty rows
# ============================================================================

# Empty rows in the middle of data are skipped.


def test_empty_row_skipped_in_checks(tmp_path: Path) -> None:
    """Verify empty row skipped in checks."""
    wb = openpyxl.Workbook()
    ws = active_sheet(wb)
    ws.title = "Checks"
    ws.append(CHECKS_HEADERS)
    ws.append(
        [None, "Speed", "never_exceeds", 220, None, None, None, None],
    )
    ws.append(  # empty row
        [None, None, None, None, None, None, None, None],
    )
    ws.append(
        [None, "Voltage", "never_below", 11.5, None, None, None, None],
    )
    p = tmp_path / "gaps.xlsx"
    wb.save(str(p))
    checks = load_checks_from_excel(p)
    assert len(checks) == 2


# ============================================================================
# Adversarial-input hardening
# ============================================================================
# Symlink rejection mirrors C++ `aletheia::detail::validate_loader_path` —
# canonicalisation would FOLLOW the link and defeat the check, so callers
# passing legitimate symlinks must resolve them first.


def test_loader_rejects_symlink(tmp_path: Path) -> None:
    """A symlinked .xlsx is refused outright."""
    real = make_checks_workbook(
        tmp_path,
        [[None, "Speed", "never_exceeds", 220, None, None, None, None]],
        filename="real.xlsx",
    )
    link = tmp_path / "link.xlsx"
    try:
        link.symlink_to(real)
    except (OSError, NotImplementedError):
        pytest.skip("symlink creation not permitted on this filesystem")
    with pytest.raises(ValidationError, match="symbolic link"):
        load_checks_from_excel(link)


def test_loader_rejects_symlinked_dbc(tmp_path: Path) -> None:
    """The DBC entry point also refuses symlinks."""
    real = make_dbc_workbook(
        tmp_path,
        [SYMLINK_DBC_ROW],
        filename="real_dbc.xlsx",
    )
    link = tmp_path / "link_dbc.xlsx"
    try:
        link.symlink_to(real)
    except (OSError, NotImplementedError):
        pytest.skip("symlink creation not permitted on this filesystem")
    with pytest.raises(ValidationError, match="symbolic link"):
        load_dbc_from_excel(link)
