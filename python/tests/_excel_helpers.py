# SPDX-FileCopyrightText: 2025 Nicolas Pelletier
# SPDX-License-Identifier: BSD-2-Clause
"""Workbook builders and wide-row fixtures shared by the Excel-loader tests.

``test_excel_loader.py`` constructs ``.xlsx`` workbooks from inline row
literals.  The DBC rows (14-16 cells) and When-Then rows (11 cells) are far
too wide to fit on one line — left inline they bloat the test module past
pylint's 1000-line cap (C0302) the moment a formatter explodes them to one
cell per line.  Factoring each distinct wide row into a named module-level
constant (mirroring the ``_dbc_helpers`` / ``_canonical_dbc`` underscore-
module / public-name convention) keeps the test module compact and the row
contents readable at their call sites.
"""

from __future__ import annotations

from typing import TYPE_CHECKING

import openpyxl  # type: ignore[import-untyped]

from aletheia.excel_loader import (
    CHECKS_HEADERS,
    DBC_HEADERS,
    WHEN_THEN_HEADERS,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.workbook import Workbook  # type: ignore[import-untyped]
    from openpyxl.worksheet.worksheet import (  # type: ignore[import-untyped]
        Worksheet,
    )

    from aletheia.excel_loader import CellValue


def active_sheet(wb: Workbook) -> Worksheet:
    """Return ``wb.active`` narrowed to ``Worksheet``.

    openpyxl types ``Workbook.active`` as ``Worksheet | None`` for the
    edge case of a workbook with zero sheets — but a freshly-constructed
    ``Workbook()`` always has its default sheet present.  Asserting that
    invariant here drops 21 ``# type: ignore[union-attr]`` suppressions
    (per ``feedback_no_suppression_without_approval.md``) at every fixture
    site.
    """
    ws: Worksheet | None = wb.active
    if ws is None:
        msg = "workbook has no active sheet — fixture broken"
        raise AssertionError(msg)
    return ws


def _as_text_cells(row: list[CellValue]) -> list[CellValue]:
    """Render every numeric (non-bool) cell as its text form (the all-text contract).

    A spreadsheet number cell stores a lossy IEEE-754 double, so the loaders require
    numeric fields to be TEXT-formatted; the ``make_*`` builders therefore write
    ``220`` / ``0.25`` as ``"220"`` / ``"0.25"`` so a row of plain Python literals
    produces an all-text sheet, exactly like the shared ``demo_workbook.xlsx``
    fixture.  Booleans (the ``Signed`` column) and strings pass through untouched.
    Tests that must produce a genuine *number* cell (to verify it is rejected) use
    :func:`make_number_cell_workbook` instead.
    """
    return [str(c) if isinstance(c, (int, float)) and not isinstance(c, bool) else c for c in row]


def make_checks_workbook(
    tmp_path: Path,
    rows: list[list[CellValue]],
    filename: str = "test.xlsx",
) -> Path:
    """Shortcut: workbook with only a Checks sheet (numerics written as text)."""
    wb = openpyxl.Workbook()
    ws = active_sheet(wb)
    ws.title = "Checks"
    ws.append(CHECKS_HEADERS)
    for row in rows:
        ws.append(_as_text_cells(row))
    p = tmp_path / filename
    wb.save(str(p))
    return p


def make_when_then_workbook(
    tmp_path: Path,
    rows: list[list[CellValue]],
    filename: str = "test.xlsx",
) -> Path:
    """Shortcut: workbook with only a When-Then sheet (numerics written as text)."""
    wb = openpyxl.Workbook()
    ws = active_sheet(wb)
    ws.title = "When-Then"
    ws.append(WHEN_THEN_HEADERS)
    for row in rows:
        ws.append(_as_text_cells(row))
    p = tmp_path / filename
    wb.save(str(p))
    return p


def make_dbc_workbook(
    tmp_path: Path,
    rows: list[list[CellValue]],
    filename: str = "test.xlsx",
) -> Path:
    """Shortcut: workbook with only a DBC sheet (numerics written as text)."""
    wb = openpyxl.Workbook()
    ws = active_sheet(wb)
    ws.title = "DBC"
    ws.append(DBC_HEADERS)
    for row in rows:
        ws.append(_as_text_cells(row))
    p = tmp_path / filename
    wb.save(str(p))
    return p


def make_number_cell_workbook(
    tmp_path: Path,
    title: str,
    headers: list[str],
    row: list[CellValue],
    filename: str = "test.xlsx",
) -> Path:
    """Build a one-row workbook writing cells with their NATIVE types.

    A numeric literal becomes a real *number* cell (a lossy IEEE-754 double).  The
    all-text-contract rejection tests need a genuine number cell to verify it is
    refused; the stringifying ``make_*`` builders deliberately cannot produce one.
    Author the row mostly as text and leave only the cell under test as a number
    literal, so the rejection fires on that field rather than on whichever numeric
    column the loader happens to read first.
    """
    wb = openpyxl.Workbook()
    ws = active_sheet(wb)
    ws.title = title
    ws.append(headers)
    ws.append(row)
    p = tmp_path / filename
    wb.save(str(p))
    return p


# ============================================================================
# DBC rows — Columns:
#   id, name, extended, dlc, signal, startbit, length, byteorder,
#   signed, factor, offset, min, max, unit[, multiplexor, multiplex value]
# ============================================================================

ENGINE_RPM_ROW: list[CellValue] = [
    256,
    "EngineData",
    None,
    8,
    "RPM",
    0,
    16,
    "little_endian",
    False,
    "0.25",
    0,
    0,
    "16383.75",
    "rpm",
]
ENGINE_TEMP_ROW: list[CellValue] = [
    256,
    "EngineData",
    None,
    8,
    "Temp",
    16,
    8,
    "little_endian",
    False,
    1,
    -40,
    -40,
    215,
    "C",
]
BRAKE_PRESSURE_ROW: list[CellValue] = [
    512,
    "BrakeData",
    None,
    4,
    "BrakePressure",
    0,
    16,
    "big_endian",
    False,
    "0.1",
    0,
    0,
    "6553.5",
    "bar",
]
ENGINE_RPM_HEX_ID_ROW: list[CellValue] = [
    "0x100",
    "EngineData",
    None,
    8,
    "RPM",
    0,
    16,
    "little_endian",
    False,
    "0.25",
    0,
    0,
    "16383.75",
    "rpm",
]
ENGINE_TEMP_SIGNED_ROW: list[CellValue] = [
    256,
    "EngineData",
    None,
    8,
    "Temp",
    0,
    8,
    "little_endian",
    True,
    1,
    -40,
    -40,
    215,
    "C",
]
SIGNED_INT_ONE_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Sig",
    0,
    8,
    "little_endian",
    1,
    1,
    0,
    0,
    255,
    "",
]
SIGNED_INT_ZERO_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Sig",
    0,
    8,
    "little_endian",
    0,
    1,
    0,
    0,
    255,
    "",
]
ENGINE_RPM_NO_UNIT_ROW: list[CellValue] = [
    256,
    "EngineData",
    None,
    8,
    "RPM",
    0,
    16,
    "little_endian",
    False,
    "0.25",
    0,
    0,
    "16383.75",
    None,
]
INVALID_BYTE_ORDER_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Sig",
    0,
    16,
    "mixed_endian",
    False,
    1,
    0,
    0,
    100,
    "",
]
INVALID_MESSAGE_ID_ROW: list[CellValue] = [
    "not_a_number",
    "Msg",
    None,
    8,
    "Sig",
    0,
    16,
    "little_endian",
    False,
    1,
    0,
    0,
    100,
    "",
]
# A minimal signal spec (name / start bit / length / byte order) shared by the
# demo DBC rows — named once so tests don't re-spell it (and don't trip the
# duplicate-code detector on the identical cells).
STD_SIGNAL_CELLS: list[CellValue] = ["Sig", "0", "8", "little_endian"]

SYMLINK_DBC_ROW: list[CellValue] = [
    "256",
    "Msg",
    "8",
    *STD_SIGNAL_CELLS,
    "FALSE",
    "1",
    "0",
    "0",
    "255",
    "",
    "",
    "",
    "",
]


# ============================================================================
# DBC rows with optional Multiplexor / Multiplex Value columns
# ============================================================================

MUX_SIGNAL_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "MuxSignal",
    8,
    8,
    "little_endian",
    False,
    1,
    0,
    0,
    255,
    "",
    "Selector",
    3,
]
ALWAYS_SIGNAL_MUX_COLS_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Sig",
    0,
    16,
    "little_endian",
    False,
    1,
    0,
    0,
    100,
    "",
    None,
    None,
]
MIXED_SELECTOR_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Selector",
    0,
    8,
    "little_endian",
    False,
    1,
    0,
    0,
    255,
    "",
    None,
    None,
]
MIXED_TEMPA_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "TempA",
    8,
    8,
    "little_endian",
    False,
    1,
    -40,
    -40,
    215,
    "C",
    "Selector",
    0,
]
MIXED_TEMPB_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "TempB",
    8,
    8,
    "little_endian",
    False,
    1,
    -40,
    -40,
    215,
    "C",
    "Selector",
    1,
]
PARTIAL_MUX_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Sig",
    0,
    16,
    "little_endian",
    False,
    1,
    0,
    0,
    100,
    "",
    "Selector",
    None,
]
PARTIAL_MUX_VALUE_ROW: list[CellValue] = [
    256,
    "Msg",
    None,
    8,
    "Sig",
    0,
    16,
    "little_endian",
    False,
    1,
    0,
    0,
    100,
    "",
    None,
    3,
]


# ============================================================================
# When-Then rows — Columns:
#   name, when_sig, when_cond, when_val, then_sig, then_cond,
#   then_val, then_min, then_max, within, sev
# ============================================================================

BRAKE_RESPONSE_WT_ROW: list[CellValue] = [
    "Brake response",
    "BrakePedal",
    "exceeds",
    50,
    "BrakeLight",
    "equals",
    1,
    None,
    None,
    100,
    None,
]
BRAKE_RESPONSE_SAFETY_WT_ROW: list[CellValue] = [
    "Brake response",
    "BrakePedal",
    "exceeds",
    50,
    "BrakeLight",
    "equals",
    1,
    None,
    None,
    100,
    "safety",
]
IGNITION_RPM_WT_ROW: list[CellValue] = [
    None,
    "Ignition",
    "equals",
    1,
    "RPM",
    "exceeds",
    500,
    None,
    None,
    2000,
    None,
]
FUEL_WT_ROW: list[CellValue] = [
    None,
    "FuelLevel",
    "drops_below",
    10,
    "FuelWarning",
    "stays_between",
    None,
    1,
    1,
    50,
    None,
]
UNKNOWN_WHEN_WT_ROW: list[CellValue] = [
    None,
    "Brake",
    "bogus",
    50,
    "BrakeLight",
    "equals",
    1,
    None,
    None,
    100,
    None,
]
UNKNOWN_THEN_WT_ROW: list[CellValue] = [
    None,
    "Brake",
    "exceeds",
    50,
    "BrakeLight",
    "bogus",
    1,
    None,
    None,
    100,
    None,
]
